# ── PATCH for batch.py ───────────────────────────────────────────────────────
# Replace the existing download_template endpoint with this version
# It accepts an optional product_code query param and adds USER_LABEL columns

from fastapi import APIRouter, HTTPException
from deps import CurrentUser

router = APIRouter()        # define first
@router.get("/template")
def download_template(
    current: CurrentUser,
    product_code: str = "IND-TERM-20",
):
    """
    Returns a CSV template with required columns + any USER_LABEL columns
    defined in the product's active BASE_PREMIUM formula.
    Pass ?product_code=IND-TERM-20 to get product-specific template.
    """
    # Standard columns always present
    standard_headers = [
        "proposal_ref", "benefit_type", "applicant_ref", "product_code", "age", "gender", "state",
        "face_amount", "coverage_term_yrs", "premium_mode", "tobacco_status",
        "height_inches", "weight_lbs", "systolic_bp", "diastolic_bp",
        "diabetes_type", "heart_condition", "annual_income", "existing_coverage",
    ]

    # Sample row values
    sample_row = {
        "proposal_ref":     "",
        "benefit_type":     "BASE",
        "applicant_ref":    "APP-001",
        "product_code":     product_code,
        "premium_mode":     "ANNUAL",
        "age":              "35",
        "gender":           "MALE",
        "state":            "MH",
        "face_amount":      "1000000",
        "coverage_term_yrs":"20",
        "tobacco_status":   "NEVER",
        "height_inches":    "68",
        "weight_lbs":       "170",
        "systolic_bp":      "120",
        "diastolic_bp":     "78",
        "diabetes_type":    "NONE",
        "heart_condition":  "NONE",
        "annual_income":    "800000",
        "existing_coverage":"0",
    }

    # Get USER_LABEL columns from product formula
    user_label_cols = []
    conn, release = _jobs_db()
    try:
        import psycopg2.extras
        conn.cursor_factory = psycopg2.extras.RealDictCursor
        cur = conn.cursor()
        cur.execute(
            """
            SELECT s.user_label, s.user_value AS default_value, s.description
            FROM uw_formula_step s
            JOIN uw_formula f ON f.id = s.formula_id
            WHERE f.product_code = %s
              AND f.formula_type = 'BASE_PREMIUM'
              AND f.is_active = true
              AND f.effective_date <= CURRENT_DATE
              AND (f.expiry_date IS NULL OR f.expiry_date >= CURRENT_DATE)
              AND s.parameter_type = 'USER_LABEL'
              AND s.user_label IS NOT NULL
            ORDER BY s.seq_no
            """,
            (product_code,),
        )
        user_label_cols = [dict(r) for r in cur.fetchall()]
        cur.close()
    except Exception:
        pass  # If formula not found, just use standard columns
    finally:
        release(conn)

    # Build headers
    extra_headers = [col["user_label"] for col in user_label_cols]
    all_headers   = standard_headers + extra_headers

    # Build sample row
    for col in user_label_cols:
        sample_row[col["user_label"]] = str(col["default_value"] or "")

    # Build CSV content
    import io
    buf = io.StringIO()
    buf.write(",".join(all_headers) + "\n")
    buf.write(",".join(sample_row.get(h, "") for h in all_headers) + "\n")

    # Add comment row explaining USER_LABEL columns
    if user_label_cols:
        buf.write(
            "# USER_LABEL columns: " +
            ", ".join(f"{c['user_label']} ({c['description'] or 'formula input'})"
                      for c in user_label_cols) + "\n"
        )

    buf.seek(0)
    filename = f"riskuw_batch_template_{product_code}.csv"
    bom_buf = io.StringIO('\ufeff' + buf.getvalue())
    return StreamingResponse(
        bom_buf,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Batch Jobs CRUD ───────────────────────────────────────────────────────────
import io, uuid, csv
from fastapi import UploadFile, File, Form, BackgroundTasks, Query
from fastapi.responses import StreamingResponse
import psycopg2.extras

def _jobs_db():
    from database import get_conn, release_conn
    conn = get_conn()
    conn.cursor_factory = psycopg2.extras.RealDictCursor
    return conn, release_conn

def _fmt_job(row: dict) -> dict:
    for f in ("submitted_at","started_at","completed_at"):
        if row.get(f): row[f] = str(row[f])
    for f in ("total_records","processed_count","approved_count",
              "declined_count","referred_count","errored_count"):
        if row.get(f) is not None: row[f] = int(row[f])
    if row.get("id"): row["id"] = str(row["id"])
    return row

@router.post("/upload")
async def upload_batch(
    background_tasks: BackgroundTasks,
    current: CurrentUser,
    file: UploadFile = File(...),
    job_name: str = Query(""),
    dry_run: bool = Query(False),
    skip_product_errors: bool = Query(False),
    policy_effective_date: str = Query(""),
    policy_expire_date: str = Query(""),
    auto_assign: bool = Query(False),
    sla_hours: int = Query(48),
    ai_engine: str = Query("rules_only"),
):
    file_bytes = await file.read()
    filename   = file.filename or "batch.csv"
    job_id     = str(uuid.uuid4())

    conn, release = _jobs_db()
    try:
        cur = conn.cursor()
        # Count rows
        try:
            text = file_bytes.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
            total = len(rows)
        except Exception:
            total = 0

        # Generate job_number
        cur.execute("SELECT COALESCE(MAX(CAST(job_number AS INTEGER)), 0) + 1 AS next_num FROM batch_jobs WHERE job_number ~ '^[0-9]+$'")
        next_row = cur.fetchone()
        job_number = str((next_row["next_num"] if next_row else 1) or 1).zfill(6)

        cur.execute("""
            INSERT INTO batch_jobs (
                id, job_number, job_name, status, total_records, dry_run,
                skip_product_errors, policy_effective_date, policy_expire_date,
                input_filename, submitted_by, submitted_at
            ) VALUES (
                %s, %s, %s, 'QUEUED', %s, %s, %s, %s, %s, %s, %s, now()
            ) RETURNING id, job_number
        """, (
            job_id, job_number,
            job_name or filename,
            total, dry_run, skip_product_errors,
            policy_effective_date or None,
            policy_expire_date or None,
            filename, current.username,
        ))
        row = cur.fetchone()
        conn.commit()

        # Store file content in DB for worker
        cur.execute("""
            UPDATE batch_jobs SET file_content = %s WHERE id = %s
        """, (file_bytes.decode("utf-8-sig", errors="replace"), job_id))
        conn.commit()
        cur.close()

        # Trigger background processing
        background_tasks.add_task(
            _run_batch_job, job_id, file_bytes, filename, dry_run,
            skip_product_errors, policy_effective_date, policy_expire_date,
            current.username, ai_engine
        )

        return {
            "job_id":     str(row["id"]),
            "job_number": row.get("job_number", ""),
            "status":     "QUEUED",
            "total_records": total,
            "message":    "Batch job queued for processing",
        }
    finally:
        release(conn)


def _run_batch_job(job_id, file_bytes, filename, dry_run,
                   skip_product_errors, eff_date, exp_date, username,
                   ai_engine="rules_only"):
    """Background task to process batch job."""
    import logging as _logging
    _logger = _logging.getLogger("uw_platform")
    try:
        from services.batch_processor import process_batch_job
        process_batch_job(job_id, file_bytes, filename, username, username, ai_engine=ai_engine)
    except (ImportError, TypeError):
        _fallback_process(job_id, file_bytes, filename, dry_run,
                         skip_product_errors, username, ai_engine)
    except Exception as e:
        _logger.error(f"Batch job {job_id} failed: {e}", exc_info=True)
        conn, release = _jobs_db()
        try:
            cur = conn.cursor()
            cur.execute("""
                UPDATE batch_jobs SET status='FAILED', error_message=%s,
                completed_at=now() WHERE id=%s
            """, (str(e), job_id))
            conn.commit()
            cur.close()
        finally:
            release(conn)


def _validate_product(cur, product_code: str) -> dict | None:
    """Returns product row if valid and active, else None."""
    cur.execute(
        "SELECT product_code, product_name, is_active FROM products WHERE product_code = %s",
        (product_code,)
    )
    row = cur.fetchone()
    if not row:
        return None
    return dict(row)


def _calculate_premium(cur, product_code: str, payload: dict, result: dict) -> dict:
    """
    Calculate premium for an APPROVED case.
    Returns dict with keys:
      - premium: float or None
      - premium_note: info/warning message string
      - user_label_values: dict of {user_label: value} used in formula
    """
    out = {"premium": None, "premium_note": "", "user_label_values": {}}

    # 1. Check if formula exists for this product
    cur.execute("""
        SELECT id FROM uw_formula
        WHERE product_code = %s
          AND formula_type = 'BASE_PREMIUM'
          AND is_active = true
          AND effective_date <= CURRENT_DATE
          AND (expiry_date IS NULL OR expiry_date >= CURRENT_DATE)
        ORDER BY effective_date DESC LIMIT 1
    """, (product_code,))
    formula_row = cur.fetchone()
    if not formula_row:
        out["premium_note"] = "Premium not calculated — no formula attached to product"
        return out

    formula_id = dict(formula_row)["id"]

    # 2. Get formula steps ordered by seq_no
    cur.execute("""
        SELECT seq_no, operator, factor, parameter_type,
               user_value, user_label, scale_id, description
        FROM uw_formula_step
        WHERE formula_id = %s
        ORDER BY seq_no
    """, (formula_id,))
    steps = [dict(r) for r in cur.fetchall()]
    if not steps:
        out["premium_note"] = "Premium not calculated — formula has no steps defined"
        return out

    # 3. Evaluate steps
    running = 0.0
    missing_params = []

    for step in steps:
        ptype    = step["parameter_type"]
        operator = step["operator"]
        factor   = float(step["factor"] or 1)

        # Resolve the step value
        step_val = None

        if ptype == "USER_VALUE":
            step_val = float(step["user_value"] or 0)

        elif ptype == "USER_LABEL":
            label_key = step["user_label"]
            raw = payload.get(label_key) or payload.get(label_key.lower())
            if raw in (None, ""):
                missing_params.append(label_key)
                continue
            try:
                step_val = float(raw)
                out["user_label_values"][label_key] = step_val
            except (ValueError, TypeError):
                missing_params.append(label_key)
                continue

        elif ptype in ("SUM_ASSURED", "FACE_AMOUNT"):
            step_val = float(payload.get("face_amount") or 0)

        elif ptype == "AGE":
            step_val = float(payload.get("age") or 0)

        elif ptype == "ANNUAL_INCOME":
            step_val = float(payload.get("annual_income") or 0)

        elif ptype == "POLICY_TERM":
            step_val = float(payload.get("coverage_term_yrs") or 0)

        elif ptype == "DEBIT_POINTS":
            step_val = float(result.get("net_debit_points") or 0)

        elif ptype == "PREVIOUS_RESULT":
            step_val = running

        elif ptype == "RATE_SCALE":
            # Rate scale lookup — use scale_id to get rate for age/term
            scale_id = step.get("scale_id")
            if not scale_id:
                missing_params.append(f"rate_scale(step {step['seq_no']})")
                continue
            try:
                age  = int(payload.get("age") or 0)
                term = int(payload.get("coverage_term_yrs") or 0)
                cur.execute("""
                    SELECT rate FROM uw_rate_scale
                    WHERE id = %s AND age = %s AND term = %s
                    LIMIT 1
                """, (scale_id, age, term))
                rate_row = cur.fetchone()
                if rate_row:
                    step_val = float(dict(rate_row)["rate"])
                else:
                    missing_params.append(f"rate_scale(age={age},term={term})")
                    continue
            except Exception:
                missing_params.append(f"rate_scale(step {step['seq_no']})")
                continue
        else:
            continue  # Unknown parameter type — skip

        # Apply operator
        val = step_val * factor
        if operator == "+":   running += val
        elif operator == "-": running -= val
        elif operator == "*": running *= val
        elif operator == "/" and val != 0: running /= val
        elif operator == "%": running += running * (val / 100)

    # 4. Build result
    if missing_params:
        out["premium_note"] = (
            f"Premium not calculated — missing parameters: {', '.join(missing_params)}"
        )
    else:
        out["premium"] = round(running, 2)
        out["premium_note"] = ""

    return out


def _get_active_user_labels(cur) -> list[dict]:
    """Returns all active user labels sorted by sort_order."""
    try:
        cur.execute("""
            SELECT label_key, label_name, data_type
            FROM system_user_label
            WHERE is_active = true
              AND effective_date <= CURRENT_DATE
              AND (expiry_date IS NULL OR expiry_date >= CURRENT_DATE)
            ORDER BY sort_order, label_name
        """)
        return [dict(r) for r in cur.fetchall()]
    except Exception:
        return []


def _fallback_process(job_id, file_bytes, filename, dry_run,
                      skip_product_errors, username, ai_engine="rules_only"):
    """Fallback batch processor if service not available."""
    import csv, io as _io, json
    import logging as _logging
    _logger = _logging.getLogger("uw_platform")
    from services.uw_engine import run_evaluation

    conn, release = _jobs_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            UPDATE batch_jobs SET status='RUNNING', started_at=now()
            WHERE id=%s
        """, (job_id,))
        conn.commit()

        text   = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(_io.StringIO(text))
        rows   = list(reader)

        # Cache valid products to avoid repeated DB hits
        cur.execute("SELECT product_code FROM products WHERE is_active = true")
        valid_products = {dict(r)["product_code"] for r in cur.fetchall()}

        approved = declined = referred = errored = 0

        # Delete existing records so re-runs get fresh AI data
        cur.execute("DELETE FROM batch_job_records WHERE job_id = %s", (job_id,))
        conn.commit()

        # ── Group rows by proposal_ref for multi-benefit processing ──────────
        # Rows with proposal_ref + non-BASE benefit_type are rider lines
        # Rows without proposal_ref OR with benefit_type=BASE only → single benefit
        from collections import OrderedDict
        proposal_groups: OrderedDict = OrderedDict()
        single_rows = []  # (original_index, row)

        for i, row in enumerate(rows, 1):
            p = {k.strip().lower(): (v.strip() if v else '') for k, v in row.items()}
            pref = p.get("proposal_ref", "").strip()
            btype = p.get("benefit_type", "").strip().upper()
            if pref:
                if pref not in proposal_groups:
                    proposal_groups[pref] = []
                proposal_groups[pref].append((i, p))
            else:
                single_rows.append((i, p))

        # Separate true multi-benefit proposals from single-benefit ones
        # (proposals with only one BASE row are treated as single-benefit)
        multi_proposals = {}
        for pref, prows in proposal_groups.items():
            if len(prows) > 1:
                multi_proposals[pref] = prows
            else:
                # Single row with proposal_ref — treat as single benefit
                single_rows.append(prows[0])

        # Process multi-benefit proposals first
        for pref, prows in multi_proposals.items():
            try:
                base_row = next((p for _, p in prows if p.get("benefit_type","").upper() == "BASE"), None)
                if not base_row:
                    base_row = prows[0][1]  # fallback to first row
                row_num = prows[0][0]

                # Build benefits list
                benefits = []
                for _, p in prows:
                    btype = p.get("benefit_type","BASE").upper() or "BASE"
                    pc = p.get("product_code","").upper()
                    if pc not in valid_products:
                        continue
                    try:
                        fa = float(p.get("face_amount",0) or 0)
                        term = int(float(p.get("coverage_term_yrs",20) or 20))
                    except:
                        fa, term = 0, 20
                    benefits.append({
                        "benefit_type": btype,
                        "product_code": pc,
                        "face_amount": fa,
                        "coverage_term_yrs": term,
                        "premium_mode": base_row.get("premium_mode","ANNUAL") or "ANNUAL",
                    })

                if not benefits:
                    errored += 1
                    continue

                # Build shared medical payload from base row
                proposal_payload = dict(base_row)
                proposal_payload["proposal_ref"] = pref
                proposal_payload["benefits"] = benefits
                # Type coercions on base row
                for int_f in ("age","coverage_term_yrs"):
                    if proposal_payload.get(int_f):
                        try: proposal_payload[int_f] = int(float(proposal_payload[int_f]))
                        except: proposal_payload.pop(int_f, None)
                for float_f in ("face_amount","a1c","annual_income"):
                    if proposal_payload.get(float_f):
                        try: proposal_payload[float_f] = float(proposal_payload[float_f])
                        except: proposal_payload.pop(float_f, None)
                for bool_f in ("hiv_positive","cirrhosis","stroke_history","kidney_disease"):
                    if proposal_payload.get(bool_f):
                        proposal_payload[bool_f] = str(proposal_payload[bool_f]).lower() in ("true","1","yes")

                if not dry_run:
                    # Call evaluate_proposal logic directly
                    from routers.underwriting import ProposalRequest, evaluate_proposal
                    import types
                    fake_auth = types.SimpleNamespace(
                        username=username, role="underwriter",
                        tenant_id="00000000-0000-0000-0000-000000000001"
                    )
                    try:
                        prop_body = ProposalRequest(**proposal_payload)
                        prop_result = evaluate_proposal(prop_body, fake_auth)
                    except Exception as pe:
                        _logger.warning(f"Proposal eval failed for {pref}: {pe}")
                        errored += len(benefits)
                        continue

                    # Write one batch_job_record per benefit line
                    for br in prop_result.get("benefits", []):
                        b_outcome = br.get("outcome","ERROR")
                        if "APPROVED" in b_outcome: approved += 1
                        elif "DECLIN" in b_outcome: declined += 1
                        elif "REFER" in b_outcome:  referred += 1
                        else:                       errored  += 1

                        cur.execute("""
                            INSERT INTO batch_job_records
                                (job_id, row_number, applicant_ref, product_code,
                                 status, outcome, risk_class, net_debit_points,
                                 primary_reason, processing_ms, created_at)
                            VALUES (%s,%s,%s,%s,'PROCESSED',%s,%s,%s,%s,%s,now())
                        """, (
                            job_id, row_num,
                            base_row.get("applicant_ref", pref),
                            br.get("product_code",""),
                            b_outcome,
                            br.get("risk_class",""),
                            br.get("net_debit_points",0),
                            f"PROPOSAL:{pref} | BENEFIT:{br.get('benefit_type','')} | {('EXCLUSIONS:' + str(br.get('exclusions',[]))) if br.get('exclusions') else ''}",
                            br.get("processing_ms",0),
                        ))
                else:
                    # Dry run — just count as would-be approved
                    approved += len(benefits)

                conn.commit()
            except Exception as pe:
                _logger.warning(f"Multi-benefit proposal {pref} failed: {pe}", exc_info=True)
                errored += 1
                conn.rollback()

        # ── Now process single-benefit rows (existing logic) ──────────────────
        rows_to_process = single_rows
        for i, row in rows_to_process:
            try:
                payload = row if isinstance(row, dict) else {k.strip().lower(): v.strip() for k,v in row.items() if v is not None}

                # ── Requirement 1: Product validation ────────────────────────
                product_code = payload.get("product_code", "").strip().upper()
                if not product_code or product_code not in valid_products:
                    errored += 1
                    err_msg = (
                        f"PROD001 — Product '{product_code}' not found in system"
                        if product_code else "PROD001 — product_code is missing"
                    )
                    cur.execute("""
                        INSERT INTO batch_job_records
                            (job_id, row_number, applicant_ref, product_code,
                             status, outcome, risk_class, net_debit_points,
                             primary_reason, error_codes, processing_ms, created_at)
                        VALUES (%s,%s,%s,%s,'ERROR','ERROR','',0,%s,%s,0,now())
                    """, (
                        job_id, i,
                        payload.get("applicant_ref", ""),
                        product_code,
                        err_msg,
                        "PROD001",
                    ))
                    if i % 50 == 0:
                        conn.commit()
                        cur.execute("""
                            UPDATE batch_jobs SET processed_count=%s,
                            approved_count=%s, declined_count=%s,
                            referred_count=%s, errored_count=%s WHERE id=%s
                        """, (i, approved, declined, referred, errored, job_id))
                        conn.commit()
                    continue  # Skip to next row — do NOT process this record

                # Type coercions
                for int_field in ("age","coverage_term_yrs","alcohol_drinks_week"):
                    if payload.get(int_field):
                        try: payload[int_field] = int(float(payload[int_field]))
                        except: payload.pop(int_field, None)
                for float_field in ("face_amount","a1c","bmi","annual_income","existing_coverage"):
                    if payload.get(float_field):
                        try: payload[float_field] = float(payload[float_field])
                        except: payload.pop(float_field, None)
                for bool_field in ("hiv_positive","cirrhosis","stroke_history",
                                   "kidney_disease","epilepsy","copd","hazardous_activity"):
                    if payload.get(bool_field):
                        payload[bool_field] = str(payload[bool_field]).lower() in ("true","1","yes")

                if not dry_run:
                    result = run_evaluation(payload, username, None)
                    outcome = result.get("outcome","ERROR")
                else:
                    result  = {}
                    outcome = "DRY_RUN"

                # ── Requirement 2: Premium calculation for APPROVED ───────────
                # Skip premium and AI for dry run — validation only
                premium      = None
                premium_note = "DRY RUN — premium not calculated" if dry_run else ""
                if not dry_run and "APPROVED" in outcome:
                    prem_result  = _calculate_premium(cur, product_code, payload, result)
                    premium      = prem_result["premium"]
                    premium_note = prem_result["premium_note"]

                # ── Reinsurance trigger ───────────────────────────────────────
                if not dry_run and "APPROVED" in outcome:
                    try:
                        from services.ri_trigger import check_and_trigger_reinsurance
                        check_and_trigger_reinsurance(batch_mode=True, 
                            conn=conn,
                            case_id=f"BATCH-{job_id[:8]}-{i}",
                            application_id=payload.get("applicant_ref", ""),
                            product_code=product_code,
                            face_amount=float(payload.get("face_amount") or 0),
                            approved_premium=premium,
                            applicant_ref=payload.get("applicant_ref", ""),
                            submitted_by=username,
                        )
                    except Exception as ri_err:
                        _logger.warning(f"RI trigger failed for row {i}: {ri_err}")

                # ── AI Scoring (if engine selected) ──────────────────────────
                ai_decision  = None
                ai_risk_tier = None
                ai_risk_score = None
                ai_narrative = None
                if not dry_run and ai_engine and ai_engine != "rules_only":
                    try:
                        from services.ai_score import get_ai_score, log_ai_decision
                        ai_payload = {
                            **payload,
                            "uw_outcome":      outcome,
                            "net_debit_points": result.get("net_debit_points", 0),
                            "engine":           ai_engine,
                        }
                        ai_result    = get_ai_score(ai_payload, engine=ai_engine, conn=conn)
                        if not ai_result.get("error"):
                            ai_decision   = ai_result.get("recommendation")
                            ai_risk_tier  = ai_result.get("risk_tier")
                            ai_risk_score = ai_result.get("risk_score")
                            ai_narrative  = ai_result.get("narrative")

                            # AI Audit Trail
                            log_ai_decision(
                                conn,
                                ai_result=ai_result,
                                input_payload=ai_payload,
                                source="BATCH",
                                job_id=job_id,
                                applicant_ref=payload.get("applicant_ref", ""),
                                product_code=product_code,
                                requested_by=username,
                            )
                    except Exception as ai_err:
                        _logger.warning(f"AI scoring failed for row {i}: {ai_err}")

                if "APPROVED" in outcome:   approved += 1
                elif "DECLINED" in outcome: declined += 1
                elif "REFERRED" in outcome: referred += 1
                else:                       errored  += 1

                # ── Requirement 3: Store input_data (for user label columns) ──
                input_data_json = json.dumps({
                    k: v for k, v in row.items()
                    if k and k.strip()
                })

                cur.execute("""
                    INSERT INTO batch_job_records
                        (job_id, row_number, applicant_ref, product_code,
                         status, outcome, risk_class, net_debit_points,
                         primary_reason, error_codes,
                         premium, premium_note, input_data,
                         ai_engine, ai_decision, ai_risk_tier, ai_risk_score, ai_narrative,
                         processing_ms, created_at)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,0,now())
                """, (
                    job_id, i,
                    payload.get("applicant_ref",""),
                    product_code,
                    "DRY_RUN" if dry_run else "PROCESSED",
                    outcome,                                    # stores DRY_RUN_APPROVED etc.
                    result.get("risk_class",""),                # show predicted risk class
                    result.get("net_debit_points",0),           # show predicted debit points
                    result.get("primary_reason",""),            # show predicted reason
                    ",".join(result.get("error_codes",[]) or []),
                    None if dry_run else premium,               # no premium for dry run
                    premium_note,
                    input_data_json,
                    None if dry_run else (ai_engine if ai_engine != "rules_only" else None),
                    None if dry_run else ai_decision,           # no AI for dry run
                    None if dry_run else ai_risk_tier,
                    None if dry_run else ai_risk_score,
                    None if dry_run else ai_narrative,
                ))
                if i % 50 == 0:
                    conn.commit()
                    cur.execute("""
                        UPDATE batch_jobs SET processed_count=%s,
                        approved_count=%s, declined_count=%s,
                        referred_count=%s, errored_count=%s
                        WHERE id=%s
                    """, (i, approved, declined, referred, errored, job_id))
                    conn.commit()

            except Exception as row_err:
                errored += 1
                _logger.warning(f"Row {i} error: {row_err}")

        conn.commit()
        cur.execute("""
            UPDATE batch_jobs SET
                status=%s, processed_count=%s, approved_count=%s,
                declined_count=%s, referred_count=%s, errored_count=%s,
                completed_at=now()
            WHERE id=%s
        """, (
            "DRY_RUN_COMPLETE" if dry_run else "COMPLETED",
            len(rows), approved, declined, referred, errored, job_id
        ))
        conn.commit()
        cur.close()
    except Exception as e:
        import traceback
        err_msg = f"Fallback batch failed: {e}\n{traceback.format_exc()}"
        try:
            _logger.error(err_msg)
        except:
            print(err_msg)
        try:
            cur.execute("""
                UPDATE batch_jobs SET status='FAILED', error_message=%s,
                completed_at=now() WHERE id=%s
            """, (str(e), job_id))
            conn.commit()
        except: pass
    finally:
        release(conn)


@router.get("/jobs")
def list_jobs(current: CurrentUser, limit: int = 50):
    conn, release = _jobs_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT id, job_number, job_name, status, total_records,
                   processed_count, approved_count, declined_count,
                   referred_count, errored_count, dry_run,
                   input_filename, error_message, submitted_by,
                   submitted_at, started_at, completed_at
            FROM batch_jobs
            ORDER BY submitted_at DESC LIMIT %s
        """, (limit,))
        rows = [_fmt_job(dict(r)) for r in cur.fetchall()]
        cur.close()
        return rows
    finally:
        release(conn)


@router.get("/jobs/{job_id}")
def get_job(job_id: str, current: CurrentUser):
    conn, release = _jobs_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT id, job_number, job_name, status, total_records,
                   processed_count, approved_count, declined_count,
                   referred_count, errored_count, dry_run,
                   input_filename, error_message, submitted_by,
                   submitted_at, started_at, completed_at
            FROM batch_jobs WHERE id=%s
        """, (job_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Job not found")

        # Get records
        cur.execute("""
            SELECT row_number, applicant_ref, product_code,
                   status, outcome, risk_class, net_debit_points,
                   primary_reason, error_codes, processing_ms
            FROM batch_job_records WHERE job_id=%s
            ORDER BY row_number LIMIT 500
        """, (job_id,))
        records = [dict(r) for r in cur.fetchall()]
        cur.close()

        result = _fmt_job(dict(row))
        result["records"] = records
        return result
    finally:
        release(conn)


@router.post("/jobs/{job_id}/cancel")
def cancel_job(job_id: str, current: CurrentUser):
    conn, release = _jobs_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            UPDATE batch_jobs SET status='CANCELLED', completed_at=now()
            WHERE id=%s AND status IN ('QUEUED','RUNNING')
        """, (job_id,))
        conn.commit()
        cur.close()
        return {"message": "Job cancelled"}
    finally:
        release(conn)


@router.get("/jobs/{job_id}/records")
def get_job_records(
    job_id: str,
    page: int = Query(default=1, ge=1),
    per_page: int = Query(default=50, le=200),
    outcome: str = Query(default=""),
    current: CurrentUser = None,
):
    """Get individual records for a batch job with pagination."""
    db, release = _jobs_db()
    try:
        import psycopg2.extras
        # cursor_factory set per-cursor instead (safe for shared connections)
        import psycopg2.extras
        cur = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        offset = (page - 1) * per_page
        outcome_filter = ""
        params = [job_id]
        if outcome:
            outcome_filter = "AND outcome ILIKE %s"
            params.append(f"%{outcome}%")

        cur.execute(f"""
            SELECT row_number, applicant_ref, product_code, status, outcome,
                   risk_class, net_debit_points, primary_reason, error_codes,
                   premium, premium_note, ai_decision, ai_risk_score
            FROM batch_job_records
            WHERE job_id = %s {outcome_filter}
            ORDER BY row_number
            LIMIT %s OFFSET %s
        """, params + [per_page, offset])
        rows = [dict(r) if hasattr(r, 'keys') else {
            'row_number': r[0], 'applicant_ref': r[1], 'product_code': r[2],
            'status': r[3], 'outcome': r[4], 'risk_class': r[5],
            'net_debit_points': r[6], 'primary_reason': r[7], 'error_codes': r[8],
            'premium': float(r[9]) if r[9] else None, 'premium_note': r[10],
            'ai_decision': r[11], 'ai_risk_score': r[12],
        } for r in cur.fetchall()]

        cur.execute("SELECT COUNT(*) FROM batch_job_records WHERE job_id = %s", (job_id,))
        count_row = cur.fetchone()
        total = count_row["count"] if count_row else 0
        cur.close()
        return {"records": rows, "total": int(total), "page": page, "per_page": per_page}
    finally:
        release(db)


@router.get('/jobs/{job_id}/download/{type}')
@router.get("/jobs/{job_id}/download/{type}")
def download_results(job_id: str, type: str, fmt: str = "csv", current: CurrentUser = None):
    import json as _json
    conn, release = _jobs_db()
    try:
        cur = conn.cursor()

        # ── Fetch records including premium, input_data and AI columns ──────────
        cur.execute("""
            SELECT r.row_number, r.applicant_ref, r.product_code,
                   r.status, r.outcome, r.risk_class, r.net_debit_points,
                   r.primary_reason, r.error_codes, r.processing_ms,
                   r.premium, r.premium_note, r.input_data,
                   r.ai_engine, r.ai_decision, r.ai_risk_tier,
                   r.ai_risk_score, r.ai_narrative
            FROM batch_job_records r
            WHERE r.job_id = %s
            ORDER BY r.row_number
        """, (job_id,))
        rows = [dict(r) for r in cur.fetchall()]

        # ── Get active user labels to include as columns ──────────────────────
        user_labels = _get_active_user_labels(cur)
        cur.close()

        # ── Parse input_data JSON and merge user label values into each row ───
        for row in rows:
            input_data = {}
            if row.get("input_data"):
                try:
                    input_data = _json.loads(row["input_data"])
                    # Normalise keys
                    input_data = {k.strip().lower(): v for k, v in input_data.items()}
                except Exception:
                    pass
            row["_input"] = input_data
            # Populate user label columns
            for ul in user_labels:
                col = ul["label_key"]
                row[f"ul_{col}"] = input_data.get(col, "")

        # ── Filter by download type ───────────────────────────────────────────
        if type == "errors":
            rows = [r for r in rows if (
                r.get("status") == "ERROR"
                or r.get("outcome") in (None, "", "ERROR", "PRODUCT_NOT_FOUND")
                or r.get("error_codes") not in (None, "")
            )]
        elif type == "summary":
            from collections import Counter
            counts = Counter(r.get("outcome") or "UNKNOWN" for r in rows)
            summary_rows = [{"outcome": k, "count": v} for k, v in counts.items()]
            export_cols  = ["outcome", "count"]

            if fmt == "xlsx":
                import openpyxl, io as _io
                wb = openpyxl.Workbook(); ws = wb.active
                ws.append(export_cols)
                for row in summary_rows:
                    ws.append([row.get(c, "") for c in export_cols])
                buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
                return StreamingResponse(buf,
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment; filename=batch_summary_{job_id[:8]}.xlsx"})
            else:
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerow(export_cols)
                for row in summary_rows:
                    writer.writerow([row.get(c, "") for c in export_cols])
                output.seek(0)
                return StreamingResponse(iter(['\ufeff' + output.getvalue()]), media_type="text/csv",
                    headers={"Content-Disposition": f"attachment; filename=batch_summary_{job_id[:8]}.csv"})

        # ── Build export columns for results / errors ─────────────────────────
        base_cols = [
            "row_number", "applicant_ref", "product_code",
            "outcome", "risk_class", "net_debit_points",
            "primary_reason", "error_codes",
            "premium", "premium_note",
            "ai_engine", "ai_decision", "ai_risk_tier", "ai_risk_score", "ai_narrative",
        ]
        # Add one column per active user label
        ul_cols = [f"ul_{ul['label_key']}" for ul in user_labels]
        # Friendly header names: label_name instead of ul_label_key
        ul_headers = [ul["label_name"] for ul in user_labels]

        export_cols    = base_cols + ul_cols
        display_headers = [
            "Row", "Applicant Ref", "Product Code",
            "Outcome", "Risk Class", "Net Debit Points",
            "Primary Reason", "Error Codes",
            "Premium (₹)", "Premium Note",
            "AI Engine", "AI Decision", "AI Risk Tier", "AI Risk Score", "AI Narrative",
        ] + ul_headers

        # ── Render xlsx or csv ────────────────────────────────────────────────
        if fmt == "xlsx":
            import openpyxl, io as _io
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = type.capitalize()

            # Header row with formatting
            ws.append(display_headers)
            for cell in ws[1]:
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = PatternFill("solid", fgColor="1a2744")
                cell.alignment = Alignment(horizontal="center")

            # Data rows with outcome colour coding
            outcome_colors = {
                "APPROVED": "d4edda", "DECLINED": "f8d7da",
                "REFERRED": "fff3cd", "ERROR": "e2e3e5",
            }
            for row in rows:
                data_row = []
                for col in export_cols:
                    val = row.get(col, "")
                    if val is None: val = ""
                    data_row.append(val)
                ws.append(data_row)
                outcome = str(row.get("outcome",""))
                color   = outcome_colors.get(outcome.split("_")[0] if "_" in outcome else outcome, "")
                if color:
                    for cell in ws[ws.max_row]:
                        cell.fill = PatternFill("solid", fgColor=color)

            # Auto-width columns
            for col_cells in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 40)

            buf = _io.BytesIO(); wb.save(buf); buf.seek(0)
            return StreamingResponse(buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=batch_{type}_{job_id[:8]}.xlsx"})
        else:
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(display_headers)
            for row in rows:
                writer.writerow([
                    "" if row.get(c) is None else row.get(c, "")
                    for c in export_cols
                ])
            output.seek(0)
            return StreamingResponse(iter(['\ufeff' + output.getvalue()]), media_type="text/csv",
                headers={"Content-Disposition": f"attachment; filename=batch_{type}_{job_id[:8]}.csv"})
    finally:
        release(conn)

@router.get("/schedules")
def list_schedules(current: CurrentUser):
    conn, release = _jobs_db()
    try:
        cur = conn.cursor()
        try:
            cur.execute("SELECT * FROM batch_recurring_schedules ORDER BY created_at DESC")
            rows = cur.fetchall()
            result = []
            for r in rows:
                row = dict(r)
                # Table column is status ('ACTIVE'/'INACTIVE'); frontend contract uses is_active.
                row["is_active"] = row.get("status") == "ACTIVE"
                for f in ("created_at", "last_run_at", "next_run_at"):
                    if row.get(f): row[f] = str(row[f])
                result.append(row)
            cur.close()
            return result
        except Exception:
            cur.close()
            return []
    finally:
        release(conn)


@router.post("/schedules")
def create_schedule(body: dict, current: CurrentUser):
    conn, release = _jobs_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO batch_recurring_schedules
                (schedule_name, cron_expression, status)
            VALUES (%s, %s, %s) RETURNING id
        """, (
            body.get("schedule_name"), body.get("cron_expression"),
            "ACTIVE" if body.get("is_active", True) else "INACTIVE",
        ))
        sid = cur.fetchone()["id"]
        conn.commit()
        cur.close()
        return {"id": sid, "message": "Schedule created"}
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        release(conn)


@router.patch("/schedules/{sid}")
def update_schedule(sid: int, body: dict, current: CurrentUser):
    conn, release = _jobs_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            UPDATE batch_recurring_schedules
            SET status=%s
            WHERE id=%s
        """, ("ACTIVE" if body.get("is_active") else "INACTIVE", sid))
        conn.commit()
        cur.close()
        return {"message": "Schedule updated"}
    finally:
        release(conn)

